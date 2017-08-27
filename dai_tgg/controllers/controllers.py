# -*- coding: utf-8 -*-
from odoo import http
from openerp import http
from openerp.http import request
from openerp.addons.web.controllers.main import serialize_exception,content_disposition
import base64
from openpyxl import load_workbook
from cStringIO import StringIO
from odoo.tools.misc import xlwt
from copy import deepcopy

def adict_flat(adict,item_seperate=';',k_v_separate = ':'):
    alist = []
    for k,v in adict.iteritems():
        if isinstance(v,dict):
            v = adict_flat(v,item_seperate=',',k_v_separate = ' ')
        alist.append(k + k_v_separate + v)
    return item_seperate.join(alist)     
        

class Binary(http.Controller):
    
    @http.route('/web/binary/download_importbdtuan' ,type='http', auth="public")
    def download_importbdtuan(self):
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('Sheet 1',cell_overwrite_ok=True)
        
        ALIGN_BORDER_dict = {'align':{'horiz': 'left','vert':'centre','wrap':'yes'},
                     "borders":{'left':'thin', 'right': 'thin', 'top': 'thin', 'bottom': 'thin'}
                     }

        title_format_dict = deepcopy(ALIGN_BORDER_dict)
        title_format_txt = adict_flat(title_format_dict)
        title_format_style = xlwt.easyxf(title_format_txt)
        worksheet.write_merge(0, 1, 0 , 0,u"Thiết bị",title_format_style)
        worksheet.write_merge(0, 1, 1 ,1,u"Hướng",title_format_style)
        
        fp = StringIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        
        return request.make_response(
            data,
            #self.from_data(columns_headers, rows),
            headers=[
                ('Content-Disposition', 'attachment; filename="import_rnas.xls"'),
                ('Content-Type', 'application/octet-stream')
            ],
            #cookies={'fileToken': token}
        )


        
    @http.route('/web/binary/download_document', type='http', auth="public")
    @serialize_exception
    def download_document(self,model,field,id,filename=None, **kw):
#         Model = request.registry[model]
#         cr, uid, context = request.cr, request.uid, request.context
#         fields = [field]
        #res = Model.read(cr, uid, [int(id)], fields, context)[0]
        win_1_linux_0 = 1
        if win_1_linux_0:
            path = 'E:\BC BAO DUONG\im\import3g.xlsx'
        else:
            PATH = 'BC BAO DUONG\im\import3g.xlsx'
            path = '/media/sf_E_DRIVE/' + PATH.replace('\\', '/')
        workbook = load_workbook(path)
        fp = StringIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return request.make_response(data,
                            [('Content-Type', 'application/octet-stream'),
                             ('Content-Disposition', content_disposition(filename))])
        
        
        Model = request.env[model]
        res = Model.browse([int(id)])
        filecontent = base64.b64decode(getattr(res,field) or '')
        #filecontent = getattr(res,field) 
        if not filecontent:
            return request.not_found()
        else:
            if not filename:
                filename = '%s_%s' % (model.replace('.', '_'), id)
            return request.make_response(filecontent,
                            [('Content-Type', 'application/octet-stream'),
                             ('Content-Disposition', content_disposition(filename))])



# class DaiTgg(http.Controller):
#     @http.route('/dai_tgg/dai_tgg/', auth='public')
#     def index(self, **kw):
#         return "Hello, world"

#     @http.route('/dai_tgg/dai_tgg/objects/', auth='public')
#     def list(self, **kw):
#         return http.request.render('dai_tgg.listing', {
#             'root': '/dai_tgg/dai_tgg',
#             'objects': http.request.env['dai_tgg.dai_tgg'].search([]),
#         })

#     @http.route('/dai_tgg/dai_tgg/objects/<model("dai_tgg.dai_tgg"):obj>/', auth='public')
#     def object(self, obj, **kw):
#         return http.request.render('dai_tgg.object', {
#             'object': obj
#         })